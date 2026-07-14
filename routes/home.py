import os
from flask import Blueprint, render_template, request, redirect, url_for
from services.workbook.manager import WorkbookManager
from services.workbook.loader import WorkbookLoader
from services.workbook.validator import WorkbookValidator
from datetime import datetime
from flask import send_file
from werkzeug.utils import secure_filename
from utils.auth import login_required
home_bp = Blueprint("home", __name__)


@home_bp.route("/")
@login_required
def home():
    manager = WorkbookManager()
    workbook = manager.get_registered_workbook()

    if workbook is None:
        return render_template("upload_master.html")

    return render_template(
        "dashboard.html",
        workbook=os.path.basename(workbook),
        errors=[],
        success=None,
        attendance_exists=False,
        unknown_morning=[],
        unknown_afternoon=[],
        summary=None,
        today=datetime.now().strftime("%A, %d %B %Y")
    )


@home_bp.route("/upload_master", methods=["GET","POST"])
@login_required
def upload_master():

    if request.method == "GET":
        return render_template("upload_master.html")

    file = request.files.get("master_workbook")

    if not file or file.filename == "":
        return render_template("upload_master.html", error="Please select a file to upload.")

    ALLOWED_EXTENSIONS = {".xlsm", ".xlsx"}
    ext = os.path.splitext(file.filename)[-1].lower()

    if ext not in ALLOWED_EXTENSIONS:
        return render_template("upload_master.html", error="Invalid file type. Please upload a .xlsm or .xlsx file.")

    destination_folder = "uploads/master"
    os.makedirs(destination_folder, exist_ok=True)
    safe_filename = secure_filename(file.filename)
    temp_path = os.path.join(destination_folder, safe_filename)
    file.save(temp_path)
    from openpyxl import load_workbook
    try:
        wb = load_workbook(temp_path, keep_vba=True)
    except Exception as e:
        os.remove(temp_path)
        return render_template("upload_master.html", error=f"Invalid workbook file: {e}")

    validator = WorkbookValidator(wb)
    result = validator.validate_workbook()

    if not result.is_valid:
        os.remove(temp_path)
        return render_template("upload_master.html", error=" | ".join(result.errors))

    manager = WorkbookManager()

    # File is already saved in uploads/master — register it directly
    # without copying (avoids Windows file lock error).
    # Store as a forward-slash path so settings.json stays
    # portable across Windows, Linux, and Mac.
    from pathlib import Path
    manager.settings["master_workbook"] = Path(temp_path).as_posix()

    import json
    with open(manager.settings_path, "w", encoding="utf-8") as f:
        json.dump(manager.settings, f, indent=4)

    return redirect(url_for("home.home"))

@home_bp.route("/download_workbook")
@login_required
def download_workbook():
    import os
    manager = WorkbookManager()
    workbook_path = manager.get_registered_workbook()

    if workbook_path is None or not os.path.exists(workbook_path):
        return redirect(url_for("home.home"))

    return send_file(
        workbook_path,
        as_attachment=True,
        download_name="Master_Workbook.xlsm"
    )