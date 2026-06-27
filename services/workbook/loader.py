from pathlib import Path

from openpyxl import load_workbook

from config.config import Config


class WorkbookLoader:
    """
    Loads the registered master workbook and
    provides access to its worksheets.
    """

    def __init__(self, workbook_path: Path):
        self.workbook_path = Path(workbook_path)
        self.workbook = None
        self.settings = Config.load_settings()

    def load_workbook(self):
        """
        Loads the workbook into memory.
        """
        self.workbook = load_workbook(self.workbook_path, keep_vba=True)
        
    def get_workbook(self):
        """
        Returns the workbook object.
        """

        return self.workbook

    def get_student_sheet(self):
        """
        Returns the Student Details worksheet.
        """

        sheet_name = self.settings["student_sheet"]
        return self.workbook[sheet_name]

    def get_attendance_sheet(self):
        """
        Returns the Attendance worksheet.
        """

        sheet_name = self.settings["attendance_sheet"]
        return self.workbook[sheet_name]